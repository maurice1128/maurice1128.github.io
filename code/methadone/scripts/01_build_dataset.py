# -*- coding: utf-8 -*-
"""
Step 01: Link urine + daily-dose data into a de-identified analysis table.

Inputs (raw data, NOT modified):
  - 2018-2022美沙冬每日劑量.xlsx  : 處方日期, 病歷號碼, 處方劑量(mg), 允許減量, 服藥劑量(cc)
  - 2018-2022驗尿結果.xlsx        : 病歷號, 看診日期(民國), 檢驗名稱(0=安非他命,1=嗎啡), 檢驗值(0陰1陽)

Outputs (scratch/analysis dir):
  - dose_cache.parquet      : tidy daily dose series [pid, date, dose_mg, dose_cc]
  - urine_morphine.parquet  : one row per morphine urine test [pid, date, positive]
  - analysis_index.parquet  : morphine tests that HAVE dose coverage, with window stats
  - id_map.csv              : 病歷號 -> pid_hash (kept local; never published)

De-identification: 病歷號 is hashed to pid_hash for any shareable artifact.
No names are ever loaded here (name column lives only in 陪同 list, handled separately).
"""
import os as _os
ROOT = _os.environ.get('MMT_DATA_ROOT', '')  # local folder holding raw data/, code/, analysis/
if not ROOT: raise SystemExit('Set MMT_DATA_ROOT before running; patient data is not distributed.')

import os, hashlib
import pandas as pd
import openpyxl

RAW  = _os.path.join(ROOT, r'raw data')
OUT  = _os.path.join(ROOT, r'analysis')
os.makedirs(OUT, exist_ok=True)

def roc_to_date(v):
    """民國整數 YYYMMDD -> pandas Timestamp. 1070115 -> 2018-01-15."""
    try:
        v = int(v)
    except (TypeError, ValueError):
        return pd.NaT
    y = v // 10000
    m = (v // 100) % 100
    d = v % 100
    if not (1 <= m <= 12 and 1 <= d <= 31):
        return pd.NaT
    try:
        return pd.Timestamp(year=y + 1911, month=m, day=d)
    except ValueError:
        return pd.NaT

def salted_hash(pid):
    return hashlib.sha256(f'methadone::{pid}'.encode()).hexdigest()[:12]

# ---- 1. dose series ----
print('loading dose (1M rows, ~1 min)...')
wb = openpyxl.load_workbook(os.path.join(RAW, '2018-2022美沙冬每日劑量.xlsx'),
                            read_only=True, data_only=True)
ws = wb['2018-2020美沙冬']
it = ws.iter_rows(values_only=True); next(it)
rows = []
for r in it:
    if r[1] is None or r[0] is None:
        continue
    rows.append((int(r[1]), pd.Timestamp(r[0]).normalize(), r[2], r[4]))
wb.close()
dose = pd.DataFrame(rows, columns=['pid', 'date', 'dose_mg', 'dose_cc'])
dose = dose.drop_duplicates(['pid', 'date']).sort_values(['pid', 'date'])
dose.to_parquet(os.path.join(OUT, 'dose_cache.parquet'))
print(f'  dose rows={len(dose)}  patients={dose.pid.nunique()}')

# ---- 2. urine morphine tests ----
print('loading urine...')
wb = openpyxl.load_workbook(os.path.join(RAW, '2018-2022驗尿結果.xlsx'),
                            read_only=True, data_only=True)
ws = wb['驗尿結果']
it = ws.iter_rows(values_only=True); next(it)
urows = []
for r in it:
    pid, dt, name, val = r[0], r[1], r[2], r[3]
    if name != 1:            # 1 = morphine only
        continue
    if val not in (0, 1):    # keep valid neg/pos only
        continue
    urows.append((int(pid), roc_to_date(dt), int(val)))
wb.close()
urine = pd.DataFrame(urows, columns=['pid', 'date', 'positive']).dropna(subset=['date'])
urine = urine.drop_duplicates(['pid', 'date']).sort_values(['pid', 'date'])
urine.to_parquet(os.path.join(OUT, 'urine_morphine.parquet'))
print(f'  morphine tests={len(urine)}  patients={urine.pid.nunique()}  '
      f'pos={int(urine.positive.sum())}  neg={int((urine.positive==0).sum())}')

# ---- 3. which tests have dose coverage ----
dose_by_pid = {pid: g.set_index('date')['dose_mg'] for pid, g in dose.groupby('pid')}
recs = []
for row in urine.itertuples(index=False):
    s = dose_by_pid.get(row.pid)
    if s is None:
        continue
    w7  = s[(s.index >= row.date - pd.Timedelta(days=7)) & (s.index <= row.date + pd.Timedelta(days=7))]
    w28 = s[(s.index >= row.date - pd.Timedelta(days=28)) & (s.index <= row.date + pd.Timedelta(days=28))]
    ref = s[s.index == row.date - pd.Timedelta(days=7)]
    recs.append((row.pid, row.date, row.positive,
                 len(w7), len(w28),
                 float(ref.iloc[0]) if len(ref) else float('nan')))
idx = pd.DataFrame(recs, columns=['pid', 'date', 'positive',
                                  'n_days_pm7', 'n_days_pm28', 'dose_day_minus7'])
idx['pid_hash'] = idx['pid'].map(salted_hash)
idx.to_parquet(os.path.join(OUT, 'analysis_index.parquet'))

# id map kept local only
pd.DataFrame({'pid': sorted(idx.pid.unique())}).assign(
    pid_hash=lambda d: d.pid.map(salted_hash)).to_csv(
    os.path.join(OUT, 'id_map.csv'), index=False)

print('\n=== ANALYSIS-READY SUMMARY ===')
print(f'morphine tests with ANY dose data : {len(idx)}  (patients={idx.pid.nunique()})')
print(f'  positive={int(idx.positive.sum())}  negative={int((idx.positive==0).sum())}')
cov7 = idx[idx.n_days_pm7 >= 8]
print(f'tests with >=8 dose days in +/-7d  : {len(cov7)}  '
      f'pos={int(cov7.positive.sum())} neg={int((cov7.positive==0).sum())}')
cov28 = idx[idx.n_days_pm28 >= 20]
print(f'tests with >=20 dose days in +/-28d: {len(cov28)}  '
      f'pos={int(cov28.positive.sum())} neg={int((cov28.positive==0).sum())}')
print('saved to', OUT)
